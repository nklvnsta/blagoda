"""
XLSX-рендерер на openpyxl.

Один файл .xlsx, по одному листу на ReportSection.
Шапка листа: title + period_label + shop_label + generated_at (в верхних строках).
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.reports.base import ReportData, ReportSection
from core.reports.renderers.formatters import format_cell

HEADER_FILL = PatternFill("solid", fgColor="F7F7F9")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
META_FONT = Font(name="Calibri", size=10, color="5C5555")
COL_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)


def render_xlsx(report: ReportData) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    for index, section in enumerate(report.sections, start=1):
        title = _unique_sheet_title(section.title or f"Лист {index}", used_titles)
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        _write_section(ws, report, section)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _unique_sheet_title(title: str, used: set[str]) -> str:
    safe = "".join(c for c in title if c not in r"[]:*?/\\")[:31] or "Лист"
    candidate = safe
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = (safe[: 31 - len(suffix)]) + suffix
        counter += 1
    return candidate


def _write_section(ws, report: ReportData, section: ReportSection) -> None:
    columns = section.columns
    last_col_letter = get_column_letter(max(len(columns), 1))

    ws.cell(row=1, column=1, value=report.title).font = TITLE_FONT
    ws.merge_cells(f"A1:{last_col_letter}1")

    meta_lines = [section.title, report.period_label]
    if report.shop_label:
        meta_lines.append(report.shop_label)
    meta_lines.append(f"Сформировано: {report.generated_at.strftime('%d.%m.%Y %H:%M')}")

    for offset, line in enumerate(meta_lines, start=2):
        cell = ws.cell(row=offset, column=1, value=line)
        cell.font = META_FONT
        ws.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=len(columns) or 1)

    if section.note:
        cell = ws.cell(row=2 + len(meta_lines), column=1, value=section.note)
        cell.font = META_FONT
        ws.merge_cells(
            start_row=2 + len(meta_lines),
            start_column=1,
            end_row=2 + len(meta_lines),
            end_column=len(columns) or 1,
        )
        header_row_idx = 3 + len(meta_lines)
    else:
        header_row_idx = 2 + len(meta_lines)

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col.title)
        cell.font = COL_HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal=col.align, vertical="center")
        if col.width:
            ws.column_dimensions[get_column_letter(col_idx)].width = col.width

    data_start = header_row_idx + 1
    for row_idx, row in enumerate(section.rows, start=data_start):
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col.key)
            text = format_cell(value, is_money=col.is_money)
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(horizontal=col.align, vertical="center")

    if section.totals:
        total_row_idx = data_start + len(section.rows)
        for col_idx, col in enumerate(columns, start=1):
            value = section.totals.get(col.key, "")
            text = format_cell(value, is_money=col.is_money)
            cell = ws.cell(row=total_row_idx, column=col_idx, value=text)
            cell.font = TOTAL_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal=col.align, vertical="center")

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)
